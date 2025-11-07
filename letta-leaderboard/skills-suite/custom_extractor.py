import json
import os
import re
from pathlib import Path
from typing import List

import anthropic
import pathspec
from letta_client import LettaMessageUnion

from letta_evals.decorators import extractor
from letta_evals.extractors.utils import (
    flatten_content,
    get_assistant_messages,
)

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd

    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


def extract_json_from_message(message: str) -> dict:
    """Extract the JSON from the message that contains 'sandbox_path' key."""
    candidates = []

    # Find all JSON within code blocks
    for match in re.finditer(r"```(?:json)?\s*\n?(.*?)```", message, re.DOTALL):
        candidates.append(match.group(1).strip())

    # Find all JSON objects in the message (look for { ... })
    for match in re.finditer(r"\{(?:[^{}]|(?:\{[^{}]*\}))*\}", message, re.DOTALL):
        candidates.append(match.group(0).strip())

    # If no candidates found, try the whole message
    if not candidates:
        candidates.append(message.strip())

    jsons_with_sandbox_path = []
    for candidate in candidates:
        try:
            parsed = json.loads(candidate)
            if isinstance(parsed, dict) and "sandbox_path" in parsed:
                jsons_with_sandbox_path.append(parsed)
                # Prefer one with 'sandbox_path' as first key
                if parsed and next(iter(parsed.keys())) == "sandbox_path":
                    return parsed
        except json.JSONDecodeError:
            continue

    if jsons_with_sandbox_path:
        return jsons_with_sandbox_path[0]

    raise ValueError(f"No JSON with 'sandbox_path' key found in message: {message}")


def json_extractor(trajectory: List[List[LettaMessageUnion]]) -> dict:
    """Extract the sandbox path from the last assistant message content."""
    messages = get_assistant_messages(trajectory)
    if not messages:
        raise ValueError("No assistant messages found")
    last_message = flatten_content(messages[-1].content).strip()
    last_message = extract_json_from_message(last_message)
    return last_message


def build_tree_structure(files: list[str]) -> str:
    """Build a tree structure from a list of file paths."""
    tree = {}

    for file_path in files:
        parts = Path(file_path).parts
        current = tree
        for part in parts:
            if part not in current:
                current[part] = {}
            current = current[part]

    def _build_tree_lines(node: dict, prefix: str = "", is_root: bool = True) -> list[str]:
        """Recursively build tree lines with proper formatting."""
        lines = []
        items = sorted(node.items())

        for i, (name, children) in enumerate(items):
            is_last_item = i == len(items) - 1

            if is_root:
                # Root level - no connector, but children get proper indentation
                connector = ""
                new_prefix = ""
                lines.append(name)
            else:
                connector = "└── " if is_last_item else "├── "
                new_prefix = prefix + ("    " if is_last_item else "│   ")
                lines.append(prefix + connector + name)

            if children:
                lines.extend(_build_tree_lines(children, new_prefix, is_root=False))

        return lines

    tree_lines = _build_tree_lines(tree)
    return "\n".join(tree_lines)


def truncate_content(content: str, max_tokens: int = 160000) -> str:
    """Truncate content if it exceeds max_tokens."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY environment variable not set")

    client = anthropic.Anthropic(api_key=api_key)

    response = client.messages.count_tokens(
        model="claude-haiku-4-5-20251001",
        messages=[{"role": "user", "content": content}],
    )

    estimated_tokens = response.input_tokens
    if estimated_tokens > max_tokens:
        # Calculate how many characters to keep (proportional to token limit)
        chars_to_keep = int(len(content) * (max_tokens / estimated_tokens))
        content = content[:chars_to_keep]
        truncation_msg = f"[WARNING: Content was truncated. Original size: ~{int(estimated_tokens):,} tokens, truncated to ~{max_tokens:,} tokens]\n\n"
        content = truncation_msg + content
        # print(f"Content truncated for judging. Original size: ~{int(estimated_tokens):,} tokens, truncated to ~{max_tokens:,} tokens")

    return content


def load_gitignore_spec(sandbox_dir: Path) -> pathspec.PathSpec:
    """Load gitignore patterns from parent directory and add default patterns."""
    # Look for .gitignore in parent directory (terminal-bench-skills/.gitignore)
    parent_gitignore = sandbox_dir.parent.parent / ".gitignore"

    patterns = []
    if parent_gitignore.exists():
        with open(parent_gitignore, "r", encoding="utf-8") as f:
            patterns.extend(f.read().splitlines())

    # Add additional patterns for files that should always be ignored
    patterns.extend(["node_modules/", "venv/"])

    return pathspec.PathSpec.from_lines("gitwildmatch", patterns)


def should_ignore_path(path: str, spec: pathspec.PathSpec, is_dir: bool = False) -> bool:
    """Check if a path should be ignored based on gitignore patterns."""
    # Add trailing slash for directory matching (gitignore convention)
    match_path = path + "/" if is_dir else path
    return spec.match_file(match_path) or (is_dir and path.startswith("."))


def read_xlsx_content(file_path: Path) -> str:
    """Read content from .xlsx file using openpyxl."""
    if not OPENPYXL_AVAILABLE:
        return "[Excel file - openpyxl not installed. Install with: pip install openpyxl]"

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        content_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content_parts.append(f"\n--- Sheet: {sheet_name} ---")

            # Get the data as rows
            for row in sheet.iter_rows(values_only=True):
                # Filter out empty rows
                if any(cell is not None for cell in row):
                    row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    content_parts.append(row_str)

        return "\n".join(content_parts)
    except Exception as e:
        return f"[Error reading .xlsx file: {str(e)}]"


def read_xls_content(file_path: Path) -> str:
    """Read content from .xls file using xlrd."""
    if not XLRD_AVAILABLE:
        return "[Excel file - xlrd not installed. Install with: pip install xlrd]"

    try:
        workbook = xlrd.open_workbook(file_path)
        content_parts = []

        for sheet in workbook.sheets():
            content_parts.append(f"\n--- Sheet: {sheet.name} ---")

            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                # Filter out empty rows
                if any(cell for cell in row):
                    row_str = " | ".join(str(cell) for cell in row)
                    content_parts.append(row_str)

        return "\n".join(content_parts)
    except Exception as e:
        return f"[Error reading .xls file: {str(e)}]"


def read_file_content(file_path: Path, relative_path: Path) -> str:
    """Read and format file content with error handling."""
    # Check if it's an Excel file first
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        content = read_xlsx_content(file_path)
        return f"=== File: {relative_path} ===\n{content}\n"
    elif suffix == ".xls":
        content = read_xls_content(file_path)
        return f"=== File: {relative_path} ===\n{content}\n"

    # Skip reading content for image files
    if suffix in [".png", ".jpg", ".jpeg", ".gif"]:
        return f"=== File: {relative_path} ===\n[Image file - content not displayed]\n"

    # For other files, try reading as text
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        return f"=== File: {relative_path} ===\n{content}\n"
    except (UnicodeDecodeError, PermissionError):
        return f"=== File: {relative_path} ===\n[Binary or unreadable file]\n"
    except Exception as e:
        return f"=== File: {relative_path} ===\n[Error reading file: {str(e)}]\n"


@extractor
def sandbox_files_extractor(trajectory: List[List[LettaMessageUnion]], config: dict) -> str:
    """Read the sandbox files from the sandbox path, respecting gitignore patterns."""
    sandbox_path = json_extractor(trajectory)["sandbox_path"]

    if not sandbox_path:
        raise ValueError("No sandbox path found")

    if not os.path.exists(sandbox_path):
        return "No sandbox files found."

    sandbox_dir = Path(sandbox_path)
    if not sandbox_dir.is_dir():
        return f"Sandbox path is not a directory: {sandbox_path}"

    # Load gitignore patterns from parent directory
    gitignore_spec = load_gitignore_spec(sandbox_dir)

    files_content = []
    file_paths = []

    for root, dirs, files in os.walk(sandbox_dir):
        root_path = Path(root)
        rel_root = root_path.relative_to(sandbox_dir)

        # Filter directories based on gitignore patterns
        dirs[:] = [
            d
            for d in dirs
            if not should_ignore_path(str(rel_root / d) if str(rel_root) != "." else d, gitignore_spec, is_dir=True)
        ]

        # Process files
        for file in sorted(files):
            file_rel_path = str(rel_root / file) if str(rel_root) != "." else file

            # Skip if file matches gitignore patterns or is hidden
            if should_ignore_path(file_rel_path, gitignore_spec) or file.startswith("."):
                continue

            file_path = root_path / file
            relative_path = file_path.relative_to(sandbox_dir.parent)
            file_paths.append(str(relative_path))

            files_content.append(read_file_content(file_path, relative_path))

    if not files_content:
        return "No readable files found in sandbox."

    tree_structure = build_tree_structure(file_paths)
    files_content_str = "".join(files_content)
    result = f"=== Sandbox Directory Structure ===\n{tree_structure}\n\n=== File Contents ===\n{files_content_str}"

    # Truncate if content is too large
    if len(result) > 160000:
        result = truncate_content(result)

    return result


@extractor
def skill_extractor(trajectory: List[List[LettaMessageUnion]], config: dict) -> str:
    """Extract the skills and additional files used from the last assistant message content."""
    message_json = json_extractor(trajectory)
    files_content = sandbox_files_extractor(trajectory, config)
    skills_content = f"Skills used: {message_json.get('skills')}\nSkills files used: {message_json.get('skills_files')}"
    return f"{skills_content}\n\nFile content:\n{files_content}"
