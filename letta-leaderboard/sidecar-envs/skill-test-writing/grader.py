#!/usr/bin/env python3
from openpyxl import load_workbook


def grade_response():
    try:
        # Load the Excel file
        wb = load_workbook("TechFlow_Revenue_Model.xlsx")
        sheet = wb.active

        score = 0.0
        feedback = []

        # Check 1: File exists and loads correctly
        try:
            wb = load_workbook("TechFlow_Revenue_Model.xlsx", data_only=True)
            feedback.append("✓ Excel file created and loads successfully")
            score += 0.2
        except Exception as e:
            feedback.append(f"✗ Failed to load Excel file: {e}")
            return score, feedback

        # Check 2: Verify formula usage (no hardcoded calculations)
        formulas_found = 0
        hardcoded_calculations = 0

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas_found += 1
                elif cell.value and isinstance(cell.value, (int, float)) and cell.row > 10:  # Skip input data rows
                    # Check if this looks like a calculated value rather than input
                    if cell.value not in [2.5, 150, 16667, 0.025, 0.08]:  # Input values
                        hardcoded_calculations += 1

        if formulas_found >= 5:  # Should have multiple formulas
            feedback.append(f"✓ Found {formulas_found} formulas (good formula usage)")
            score += 0.3
        else:
            feedback.append(f"✗ Only found {formulas_found} formulas (insufficient)")
            score += 0.1

        if hardcoded_calculations <= 2:  # Should be minimal hardcoded calculations
            feedback.append(f"✓ Minimal hardcoded calculations ({hardcoded_calculations})")
            score += 0.2
        else:
            feedback.append(f"✗ Too many hardcoded calculations ({hardcoded_calculations})")

        # Check 3: Color coding verification
        color_violations = 0
        proper_colors = 0

        for row in sheet.iter_rows():
            for cell in row:
                if cell.font and cell.font.color:
                    if hasattr(cell.font.color, "rgb"):
                        rgb = str(cell.font.color.rgb)
                        # Check for blue text (input assumptions)
                        if rgb == "FF0000FF":  # Blue
                            proper_colors += 1
                        # Check for black text (formulas)
                        elif rgb == "FF000000":  # Black
                            proper_colors += 1
                        # Check for other expected colors
                        elif rgb in ["FF008000", "FFFF0000"]:  # Green or Red
                            proper_colors += 1

                # Check for yellow background
                if cell.fill and cell.fill.start_color:
                    if hasattr(cell.fill.start_color, "rgb"):
                        if str(cell.fill.start_color.rgb) == "FFFFFF00":  # Yellow
                            proper_colors += 1

        if proper_colors > 0:
            feedback.append(f"✓ Found {proper_colors} cells with proper color coding")
            score += 0.1
        else:
            feedback.append("✗ No color coding found")

        # Check 4: Number formatting
        formatting_good = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, (int, float)):
                    if cell.number_format:
                        # Check for currency formatting
                        if "$" in str(cell.number_format):
                            formatting_good += 1
                        # Check for percentage formatting
                        elif "%" in str(cell.number_format):
                            formatting_good += 1

        if formatting_good > 0:
            feedback.append(f"✓ Found {formatting_good} cells with proper number formatting")
            score += 0.1
        else:
            feedback.append("✗ No proper number formatting found")

        # Check 5: Formula error check
        error_count = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if any(err in str(cell.value) for err in ["#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?"]):
                        error_count += 1

        if error_count == 0:
            feedback.append("✓ No formula errors found")
            score += 0.1
        else:
            feedback.append(f"✗ Found {error_count} formula errors")

        return min(score, 1.0), feedback

    except Exception as e:
        return 0.0, [f"Error during grading: {e}"]


if __name__ == "__main__":
    score, feedback = grade_response()
    print(f"Score: {score}")
    print("Feedback:")
    for item in feedback:
        print(item)
